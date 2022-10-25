import base64

from odoo import models, fields, api
from io import BytesIO
import openpyxl

from odoo.modules import get_module_resource


class SpendingReportWizard(models.TransientModel):
    _name = 'spending.report.wizard'

    month = fields.Selection([(str(i), f"Thang {i}") for i in range(1, 13)],
                             default=False, index=True, string='Thang', required=True)

    department_ids = fields.Many2many('hr.department', 'spending_report_wizard_department_rel',
                                      string='Phong ban')
    report_ids = fields.One2many('spending.report', 'wizard_id', ondelete='cascade')

    name = fields.Char(compute='_compute_name')
    file = fields.Binary('Report excel file')

    def _compute_name(self):
        for rec in self:
            rec.name = f"Bao cao chi tieu thang {rec.month}"

    # noinspection PyTypeChecker
    def show_report_btn(self):
        self.ensure_one()
        if self.department_ids:
            domain = [(5,)] + [(0, 0, {'department_id': x.id}) for x in self.department_ids]
        else:
            domain = [(5,)] + [(0, 0, {'department_id': x.id})
                               for x in self.env['hr.department'].sudo().search([])]
        self.update({
            'report_ids': domain
        })

        id_list = [x.id for x in self.report_ids]
        # id_list = []
        # for rec in self.report_ids:
        #     id_list.append(rec.id)
        view_id = self.env.ref('purchase_inherit.spending_report_tree').id
        return {
            'type': 'ir.actions.act_window',
            'name': 'show spending report',
            'view_mode': 'tree',
            'view_id': view_id,
            'res_model': 'spending.report',
            'domain': [('id', 'in', id_list)],
            'target': 'current',
        }

    @api.model
    def get_view(self):
        view_id = self.env.ref('purchase_inherit.spending_report_form').id
        # maybe get_formview_action() will be better?
        return {
            'type': 'ir.actions.act_window',
            'name': 'Show spending report form',
            'view_mode': 'form',
            'view_id': view_id,
            'res_model': 'spending.report.wizard',
            'target': 'new',
        }

    def export_excel(self):
        pass


class SpendingReport(models.TransientModel):
    _name = 'spending.report'

    wizard_id = fields.Many2one('spending.report.wizard', ondelete='cascade')
    # month = fields.Selection([(str(i), f"Thang {i}") for i in range(1, 13)],
    #                          default=False, index=True, string='Thang', required=True)
    department_id = fields.Many2one('hr.department',
                                    string='Ten phong ban')
    currency_id = fields.Many2one('res.currency', related='department_id.currency_id')
    spending_limit = fields.Monetary(related='department_id.spending_limit', currency_field='currency_id')
    spending = fields.Monetary('Chi tieu thuc te', compute='_compute_department_spending',
                               currency_field='currency_id', readonly=True)

    def export_excel(self):
        wb = openpyxl.load_workbook(
            get_module_resource('purchase_inherit', 'static/template', 'spending_report_template.xlsx'))
        ws = wb['Sheet1']
        for row in range(0, len(self)):
            ws.cell(row=row + 7, column=1).value = self[row].department_id.name
            ws.cell(row=row + 7, column=2).value = self[row].spending_limit
            ws.cell(row=row + 7, column=3).value = self[row].spending
        content = BytesIO()
        wb.save(content)
        out = base64.encodebytes(content.getvalue())
        self.wizard_id.file = out

        view_id = self.env.ref('purchase_inherit.report_excel_exported').id

        return {
            'type': 'ir.actions.act_window',
            'name': 'Excel export',
            'view_mode': 'form',
            'view_id': view_id,
            'res_model': 'spending.report.wizard',
            'res_id': self.wizard_id.id,
            'target': 'new',
            'flags': {'mode': 'readonly'},
        }

    @api.depends('department_id')
    def _compute_department_spending(self):
        for rec in self:
            def is_correct_order(search_rec):
                return search_rec.department_id == rec.department_id \
                       and search_rec.date_order.month == int(rec.wizard_id.month)

            orders = self.env['purchase.order'].sudo().search(
                [('department_id', '!=', False), ('state', '=', 'purchase')]
            ).filtered(is_correct_order)
            rec.spending = sum([x.amount_total for x in orders])
