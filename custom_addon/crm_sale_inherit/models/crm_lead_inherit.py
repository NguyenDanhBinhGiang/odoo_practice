import base64
from copy import copy
from io import BytesIO
from typing import Iterable

# noinspection PyPackageRequirements
import openpyxl

import odoo.exceptions
from odoo import models, fields, api
from odoo.modules import get_module_resource


class CrmLeadInherit(models.Model):
    _inherit = 'crm.lead'
    _sql_constraints = [('min_revenue_check', 'CHECK(min_revenue>0)', 'Doanh thu toi thieu phai >0')]

    min_revenue = fields.Monetary('Doanh thu toi thieu', currency_field='company_currency')
    month_open = fields.Integer(compute='_compute_month_opened', store=True)

    @api.depends('date_open')
    def _compute_month_opened(self):
        for rec in self:
            if rec.date_open:
                rec.month_open = rec.date_open.month
            else:
                rec.month_open = False

    # noinspection PyUnresolvedReferences
    @api.constrains('user_id')
    def assign_constraint(self):
        if not self.user_id.sale_team_id == self.env.user.sale_team_id:
            if not self.user_has_groups('crm_sale_inherit.sale_department_manager'):
                raise odoo.exceptions.UserError('You can only assign lead to other member of your sale team')

    # noinspection PyUnresolvedReferences
    def action_set_lost(self, **kwargs):
        if self.priority == '3':
            if not self.user_has_groups('crm_sale_inherit.sale_department_manager'):
                raise odoo.exceptions.UserError('Chi truong phong moi co the mark lost khi priority la cao nhat')

        super(CrmLeadInherit, self).action_set_lost(**kwargs)

    @staticmethod
    def sorted_unique(in_arr: Iterable) -> list:
        """
        Return a sorted list of unique elements in an Iterable
        @param in_arr: Iterable
        @return: list
        """
        result = list(set(in_arr))
        result.sort()
        return result

    def export_excel(self):
        records = self.search(self.env.context['active_domain'])
        teams = records.mapped('team_id')

        # Group data
        # group by team
        data = {team: [rec for rec in records if rec.team_id == team] for team in teams}
        for team, record_set in data.items():
            # group by sale person
            people = self.sorted_unique([x.user_id for x in record_set])
            group_by_person = {person: [x for x in record_set if x.user_id == person] for person in people}
            data[team] = group_by_person

        # export datas to excel
        out = self.fill_excel(data)
        self.env['file.export'].sudo().search([]).unlink()
        file_output = self.env['file.export'].sudo().create({
            'file': out,
            'file_name': "Lead Opportunity Report.xlsx"
        })
        return file_output.show_pop_up()

    # noinspection PyProtectedMember
    @staticmethod
    def fill_excel(data: dict) -> bytes:
        """
        fill data to an excel file
        @param data: and dic of grouped data
        @return: excel file
        """
        # export datas to excel
        wb = openpyxl.load_workbook(get_module_resource('crm_sale_inherit', 'static/template',
                                                        'lead_report_template.xlsx'))
        ws = wb['Sheet1']
        row = 5
        sum_rows = []
        for team in data.keys():
            team_start_row = row + 1
            for person in data[team]:
                person_start_row = row + 1
                for rec in data[team][person]:
                    row += 1
                    ws.cell(row=row, column=1).value = rec.name
                    ws.cell(row=row, column=1)._style = copy(ws.cell(row=6, column=1)._style)

                    ws.cell(row=row, column=2).value = person.name
                    ws.cell(row=row, column=2)._style = copy(ws.cell(row=6, column=1)._style)

                    ws.cell(row=row, column=3).value = team.name
                    ws.cell(row=row, column=3)._style = copy(ws.cell(row=6, column=1)._style)

                    ws.cell(row=row, column=4).value = rec.min_revenue
                    ws.cell(row=row, column=4)._style = copy(ws.cell(row=6, column=1)._style)

                    ws.cell(row=row, column=5).value = rec.sale_amount_total
                    ws.cell(row=row, column=5)._style = copy(ws.cell(row=6, column=1)._style)

                # merge person
                ws.merge_cells(start_row=person_start_row, start_column=2, end_row=row, end_column=2)
                ws.cell(row=person_start_row, column=2)._style = copy(ws.cell(row=4, column=1)._style)

            # merge team
            ws.merge_cells(start_row=team_start_row, start_column=3, end_row=row, end_column=3)
            ws.cell(row=team_start_row, column=3)._style = copy(ws.cell(row=4, column=2)._style)
            row += 1

            ws.cell(row=row, column=4).value = "Tong"
            ws.cell(row=row, column=4)._style = copy(ws.cell(row=4, column=3)._style)
            ws.cell(row=row, column=5).value = f"=SUM(E{team_start_row}:E{row - 1})"
            sum_rows.append(row)

            row += 1

        # sum of everything
        row += 1
        ws.cell(row=row, column=4).value = "Tong"
        ws.cell(row=row, column=4)._style = copy(ws.cell(row=4, column=4)._style)
        value = f'=SUM({",".join(f"E{i}" for i in sum_rows)})'
        ws.cell(row=row, column=5).value = value

        # delete sample cells
        ws.cell(row=4, column=1).value = None
        ws.cell(row=4, column=1)._style = copy(ws.cell(row=99, column=99)._style)

        ws.cell(row=4, column=2).value = None
        ws.cell(row=4, column=2)._style = copy(ws.cell(row=99, column=99)._style)

        ws.cell(row=4, column=3).value = None
        ws.cell(row=4, column=3)._style = copy(ws.cell(row=99, column=99)._style)

        ws.cell(row=4, column=4).value = None
        ws.cell(row=4, column=4)._style = copy(ws.cell(row=99, column=99)._style)

        # save it
        content = BytesIO()
        wb.save(content)
        return base64.encodebytes(content.getvalue())

    # noinspection PyShadowingNames,SpellCheckingInspection
    @api.model
    def read_group(self, domain, fields, groupby, offset=0, limit=None, orderby=False, lazy=True):
        """
        override read_group to allow computed field sale_amount_total to have sum in group by search view.
        store=True works too btw.
        """
        res = super(CrmLeadInherit, self).read_group(domain, fields, groupby, offset=offset, limit=limit,
                                                     orderby=orderby, lazy=lazy)
        if 'sale_amount_total' in fields:
            for line in res:
                if '__domain' in line:
                    lines = self.search(line['__domain'])
                    sum_total_real_income = 0.0
                    for record in lines:
                        sum_total_real_income += record.sale_amount_total
                    line['sale_amount_total'] = sum_total_real_income

        return res

    # noinspection PyUnresolvedReferences
    def write(self, vals):
        if self.quotation_count > 0 and 'min_revenue' in vals:
            raise odoo.exceptions.UserError('Ban ko the edit tiem nang khi quotation_count > 0')
        super(CrmLeadInherit, self).write(vals)
